from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\W11\Downloads\Forge Master.xlsx")
OUTPUT_DIR = Path(__file__).resolve().parents[1] / "src" / "data"


RESOURCE_MAPPING_NOTE = (
    "The workbook stores several reward tables without explicit row labels in the exported cells. "
    "V1 maps the first four relevant rows into ascension resources and preserves source row metadata "
    "so the mapping can be corrected easily if clan verification provides a better lookup."
)


@dataclass
class PillarSpec:
    name: str
    sheet_name: str
    primary_resource: str
    reserve_total_cell: str
    reserve_total_with_max_tech_cell: str
    reserve_cost_cell: str
    reserve_with_max_tech_cost_cell: str
    legendary_level_cell: str
    legendary_chance_cell: str


PILLARS = [
    PillarSpec(
        name="skills",
        sheet_name="Skills",
        primary_resource="tickets",
        reserve_total_cell="C24",
        reserve_total_with_max_tech_cell="E24",
        reserve_cost_cell="E16",
        reserve_with_max_tech_cost_cell="F16",
        legendary_level_cell="C16",
        legendary_chance_cell="D16",
    ),
    PillarSpec(
        name="pets",
        sheet_name="Pets",
        primary_resource="eggshells",
        reserve_total_cell="C25",
        reserve_total_with_max_tech_cell="E25",
        reserve_cost_cell="E17",
        reserve_with_max_tech_cost_cell="F17",
        legendary_level_cell="C17",
        legendary_chance_cell="D17",
    ),
    PillarSpec(
        name="mounts",
        sheet_name="Mount",
        primary_resource="clockwinders",
        reserve_total_cell="C26",
        reserve_total_with_max_tech_cell="E26",
        reserve_cost_cell="E18",
        reserve_with_max_tech_cost_cell="F18",
        legendary_level_cell="C18",
        legendary_chance_cell="D18",
    ),
]


def parse_short_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().lower().replace(",", "")
    if text in {"", "n/a", "none", "-"}:
        return 0.0
    multiplier = 1
    if text.endswith("k"):
        multiplier = 1_000
        text = text[:-1]
    elif text.endswith("m"):
        multiplier = 1_000_000
        text = text[:-1]
    return float(text) * multiplier


def normalize_rank_label(value: Any, number_format: str, top_rank_index: int | None = None) -> str:
    if isinstance(value, datetime) and number_format == "m-d":
        top_labels = ["1st", "2nd", "3rd"]
        if top_rank_index is not None and 0 <= top_rank_index < len(top_labels):
            return top_labels[top_rank_index]
        return f"{value.month}-{value.day}"
    return str(value).strip()


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def extract_progression_sheet(wb, spec: PillarSpec) -> dict[str, Any]:
    ws = wb[spec.sheet_name]
    rarity_headers = ["common", "rare", "epic", "legendary", "ultimate", "mythic"]
    levels = []
    total_primary_cost = 0

    for row in range(3, ws.max_row + 1):
        level = ws.cell(row, 1).value
        summons_required = ws.cell(row, 2).value
        cost_per_level = ws.cell(row, 3).value
        if not isinstance(level, (int, float)) or int(level) >= 100:
            continue

        rarity_odds = {}
        for offset, rarity in enumerate(rarity_headers, start=4):
            rarity_odds[rarity] = float(ws.cell(row, offset).value or 0)

        level_data = {
            "level": int(level),
            "summonsRequired": int(float(summons_required or 0)),
            "costPerLevel": int(float(cost_per_level or 0)),
            "rarityOdds": rarity_odds,
        }
        total_primary_cost += level_data["costPerLevel"]
        levels.append(level_data)

    return {
        "pillar": spec.name,
        "sheetName": spec.sheet_name,
        "primaryResource": spec.primary_resource,
        "levels": levels,
        "source": {
            "type": "workbook",
            "workbookPath": str(WORKBOOK_PATH),
            "sheet": spec.sheet_name,
            "headerRow": 2,
        },
        "summary": {
            "targetLevel": 100,
            "totalSummonsToLevel100": sum(level["summonsRequired"] for level in levels),
            "totalPrimaryResourceCost": total_primary_cost,
        },
    }


def get_rarity_recovery(progression: dict[str, Any], rarity_key: str, threshold: float = 5.0) -> dict[str, Any]:
    cumulative_cost = 0
    for level in progression["levels"]:
        rarity_odds = level["rarityOdds"].get(rarity_key, 0)
        if rarity_odds >= threshold:
            return {
                "targetLevel": level["level"],
                "pullChance": rarity_odds / 100,
                "reserveCost": cumulative_cost,
            }
        cumulative_cost += level["costPerLevel"]

    return {
        "targetLevel": 100,
        "pullChance": 0,
        "reserveCost": cumulative_cost,
    }


def extract_ascension_targets(wb) -> dict[str, Any]:
    ws = wb["Ascension"]
    base = {
        "gold": int(float(ws["C3"].value or 0)),
        "tickets": int(float(ws["D3"].value or 0)),
        "eggshells": int(float(ws["E3"].value or 0)),
        "clockwinders": int(float(ws["F3"].value or 0)),
    }
    maxed = {
        "gold": int(round(float(ws["C6"].value or 0))),
        "tickets": int(round(float(ws["D6"].value or 0))),
        "eggshells": int(round(float(ws["E6"].value or 0))),
        "clockwinders": int(round(float(ws["F6"].value or 0))),
    }

    progressions = {spec.name: extract_progression_sheet(wb, spec) for spec in PILLARS}
    pillar_targets: dict[str, Any] = {}
    for spec in PILLARS:
        epic_recovery = get_rarity_recovery(progressions[spec.name], "epic")
        reserve_total = int(round(float(ws[spec.reserve_total_cell].value or 0)))
        reserve_total_with_max = int(round(float(ws[spec.reserve_total_with_max_tech_cell].value or 0)))
        reserve_cost = int(round(float(ws[spec.reserve_cost_cell].value or 0)))
        reserve_with_max_cost = int(round(float(ws[spec.reserve_with_max_tech_cost_cell].value or 0)))
        legendary_level = int(round(float(ws[spec.legendary_level_cell].value or 0)))
        legendary_chance = float(ws[spec.legendary_chance_cell].value or 0)

        pillar_targets[spec.name] = {
            "primaryResource": spec.primary_resource,
            "minimumAscend": {
                spec.primary_resource: base[spec.primary_resource],
            },
            "safeAscend": {
                spec.primary_resource: base[spec.primary_resource] + epic_recovery["reserveCost"],
            },
            "optimalReset": {
                spec.primary_resource: reserve_total,
            },
            "withMaxTech": {
                spec.primary_resource: maxed[spec.primary_resource],
            },
            "safeAscendWithMaxTech": {
                spec.primary_resource: maxed[spec.primary_resource] + epic_recovery["reserveCost"],
            },
            "optimalResetWithMaxTech": {
                spec.primary_resource: reserve_total_with_max,
            },
            "epicRecovery": epic_recovery,
            "legendaryRecovery": {
                "targetLevel": legendary_level,
                "legendaryPullChance": legendary_chance,
                "reserveCost": reserve_cost,
                "reserveCostWithMaxTech": reserve_with_max_cost,
            },
        }

    return {
        "targets": {
            "baseAscension": base,
            "withMaxTech": maxed,
        },
        "pillarTargets": pillar_targets,
        "modifierFormula": {
            "effectiveRequirementMultiplier": "(1 - discountPct) / (1 + extraDropPct)",
            "effectiveFinalDiscount": "1 - ((1 - discountPct) / (1 + extraDropPct))",
            "rounding": "ceil",
            "verifiedAgainstSheet": "Discount and Extra Drop Final D",
        },
        "notes": [
            "Base totals are sourced directly from the Ascension sheet.",
            "Safe ascension uses the first sheet level where Epic chance reaches at least 5%, summed with the base ascension requirement.",
            "Optimal reset uses the legendary recovery table in rows 16-18 and 24-26, which targets a 5% legendary pull chance.",
        ],
        "source": {
            "type": "workbook",
            "workbookPath": str(WORKBOOK_PATH),
            "sheet": "Ascension",
        },
    }


def extract_forge_progression(wb) -> dict[str, Any]:
    ws = wb["Forge"]
    levels = []
    for row in range(3, ws.max_row + 1):
        level = ws.cell(row, 1).value
        total_cost = ws.cell(row, 3).value
        if not isinstance(level, (int, float)) or int(level) >= 100:
            continue
        levels.append(
            {
                "level": int(level),
                "goldCost": int(parse_short_number(total_cost)),
            }
        )

    return {
        "pillar": "forge",
        "levels": levels,
        "summary": {
            "targetLevel": 100,
            "totalGoldToLevel100": sum(level["goldCost"] for level in levels),
        },
        "source": {
            "type": "workbook",
            "workbookPath": str(WORKBOOK_PATH),
            "sheet": "Forge",
            "headerRow": 2,
        },
    }


def extract_clan_war_rewards(wb) -> dict[str, Any]:
    ws = wb["Clan War Rewards"]
    tiers = ["S", "A", "B", "C", "D", "E"]
    mapped_rows = {
        "tickets": 4,
        "gold": 5,
        "eggshells": 6,
        "clockwinders": 7,
    }
    extra_rows = {
        "unmappedRow5": 8,
        "unmappedRow6": 9,
    }

    rewards = {}
    for index, tier in enumerate(tiers):
        win_col = 2 + index * 2
        loss_col = win_col + 1
        rewards[tier] = {
            "sourceColumns": {"win": win_col, "loss": loss_col},
            "win": {resource: int(parse_short_number(ws.cell(row, win_col).value)) for resource, row in mapped_rows.items()},
            "loss": {resource: int(parse_short_number(ws.cell(row, loss_col).value)) for resource, row in mapped_rows.items()},
            "extraUnmappedRows": {
                label: {
                    "win": int(parse_short_number(ws.cell(row, win_col).value)),
                    "loss": int(parse_short_number(ws.cell(row, loss_col).value)),
                }
                for label, row in extra_rows.items()
            },
        }

    return {
        "periodDays": 7,
        "tiers": rewards,
        "assumptions": [RESOURCE_MAPPING_NOTE],
        "source": {
            "type": "workbook",
            "workbookPath": str(WORKBOOK_PATH),
            "sheet": "Clan War Rewards",
        },
    }


def extract_individual_clan_rewards(wb) -> dict[str, Any]:
    ws = wb["Individual Clan Rewards"]
    milestones = []
    for col in range(2, 15):
        label = str(ws.cell(2, col).value).strip()
        milestone_data = {
            "milestone": label,
            "sourceColumn": col,
            "rewards": {
                "tickets": int(parse_short_number(ws.cell(4, col).value)),
                "eggshells": int(parse_short_number(ws.cell(5, col).value)),
                "gold": int(parse_short_number(ws.cell(6, col).value)),
                "clockwinders": int(parse_short_number(ws.cell(8, col).value)),
            },
            "extraUnmappedRows": {
                "unmappedRow3": int(parse_short_number(ws.cell(3, col).value)),
                "unmappedRow7": int(parse_short_number(ws.cell(7, col).value)),
            },
        }
        milestones.append(milestone_data)

    return {
        "periodDays": 7,
        "milestones": milestones,
        "assumptions": [
            RESOURCE_MAPPING_NOTE,
            "The gold mapping in individual clan rewards is seeded from the row containing the only clearly large gold-like values (15k increments).",
        ],
        "source": {
            "type": "workbook",
            "workbookPath": str(WORKBOOK_PATH),
            "sheet": "Individual Clan Rewards",
        },
    }


def extract_ranked_league_rewards(wb) -> dict[str, Any]:
    ws = wb["Ranked League Rewards"]
    sections = [
        {"league": "diamond", "title_cell": "A3", "condition_cell": "A4", "rank_col": 1, "value_col_start": 2, "rows": range(5, 13)},
        {"league": "platinum", "title_cell": "I3", "condition_cell": "I4", "rank_col": 9, "value_col_start": 10, "rows": range(5, 13)},
        {"league": "gold", "title_cell": "A13", "condition_cell": "A14", "rank_col": 1, "value_col_start": 2, "rows": range(15, 23)},
        {"league": "silver", "title_cell": "I13", "condition_cell": "I14", "rank_col": 9, "value_col_start": 10, "rows": range(15, 23)},
        {"league": "bronze", "title_cell": "A23", "condition_cell": "A24", "rank_col": 1, "value_col_start": 2, "rows": range(25, 33)},
        {"league": "unranked", "title_cell": "I23", "condition_cell": "I24", "rank_col": 9, "value_col_start": 10, "rows": range(25, 33)},
    ]

    result = {"periodDays": 7, "leagues": {}, "assumptions": [RESOURCE_MAPPING_NOTE]}

    for section in sections:
        entries = []
        top_rank_index = 0
        for row in section["rows"]:
            rank_value = ws.cell(row, section["rank_col"]).value
            if rank_value is None:
                continue
            rank_label = normalize_rank_label(
                rank_value,
                ws.cell(row, section["rank_col"]).number_format,
                top_rank_index if isinstance(rank_value, datetime) else None,
            )
            if isinstance(rank_value, datetime):
                top_rank_index += 1
            entries.append(
                {
                    "rankBracket": rank_label,
                    "sourceRow": row,
                    "rewards": {
                        "tickets": int(parse_short_number(ws.cell(row, section["value_col_start"] + 2).value)),
                        "gold": int(parse_short_number(ws.cell(row, section["value_col_start"] + 1).value)),
                        "eggshells": int(parse_short_number(ws.cell(row, section["value_col_start"] + 3).value)),
                        "clockwinders": int(parse_short_number(ws.cell(row, section["value_col_start"] + 5).value)),
                    },
                    "extraUnmappedColumns": {
                        "unmappedCol1": int(parse_short_number(ws.cell(row, section["value_col_start"]).value)),
                        "unmappedCol5": int(parse_short_number(ws.cell(row, section["value_col_start"] + 4).value)),
                    },
                }
            )

        result["leagues"][section["league"]] = {
            "label": ws[section["title_cell"]].value,
            "placementRule": ws[section["condition_cell"]].value,
            "entries": entries,
        }

    return result


def build_dungeon_yields() -> dict[str, Any]:
    anchor_stage_index = 146
    mid_stage_index = 84
    keys_per_day = 2
    base_daily_tickets = 400
    base_daily_eggshells = 200
    mid_daily_tickets = 740
    mid_daily_eggshells = 284
    anchor_daily_tickets = 539 * 2
    anchor_daily_eggshells = 173 * 2
    steps_from_base = anchor_stage_index - 1

    return {
        "editable": True,
        "note": "Dungeon progression uses stage labels like 1-1 through 40-10. Eggshells use a rounded per-key linear scale anchored at 1-1, 9-4, and 15-6. Skill tickets use a rounded per-key fitted curve anchored at 1-1, 9-4, and 15-6.",
        "keysPerDay": keys_per_day,
        "stagesPerWorld": 10,
        "worlds": 40,
        "baseStage": {
            "world": 1,
            "stage": 1,
            "stageIndex": 1,
            "dailyYields": {
                "gold": 0,
                "tickets": base_daily_tickets,
                "eggshells": base_daily_eggshells,
                "clockwinders": 0,
            },
        },
        "midStage": {
            "world": 9,
            "stage": 4,
            "stageIndex": mid_stage_index,
            "dailyYields": {
                "gold": 0,
                "tickets": mid_daily_tickets,
                "eggshells": mid_daily_eggshells,
                "clockwinders": 0,
            },
        },
        "anchorStage": {
            "world": 15,
            "stage": 6,
            "stageIndex": anchor_stage_index,
            "dailyYields": {
                "gold": 0,
                "tickets": anchor_daily_tickets,
                "eggshells": anchor_daily_eggshells,
                "clockwinders": 0,
            },
        },
        "ticketFormula": {
            "kind": "quadraticPerKey",
            "coefficients": {
                "a": 0.004673197796730477,
                "b": 1.6509709583615688,
                "c": 198.3443558438417,
            },
        },
        "eggshellFormula": {
            "kind": "roundedLinearPerKey",
            "basePerKey": 100,
            "incrementPerStage": 0.505,
        },
        "perStageDailyIncrement": {
            "gold": 0,
            "tickets": 0,
            "eggshells": 0,
            "clockwinders": 0,
        },
    }


def build_app_config() -> dict[str, Any]:
    return {
        "resources": [
            {"id": "tickets", "label": "Skill Tickets", "shortLabel": "Skill Tickets", "tone": "sky"},
            {"id": "eggshells", "label": "Eggshells", "shortLabel": "Eggshells", "tone": "emerald"},
            {"id": "clockwinders", "label": "Clockwinders", "shortLabel": "Clockwinders", "tone": "rose"},
        ],
        "pillars": [
            {"id": "skills", "label": "Skills", "primaryResource": "tickets"},
            {"id": "pets", "label": "Pets", "primaryResource": "eggshells"},
            {"id": "mounts", "label": "Mount", "primaryResource": "clockwinders"},
        ],
        "targetModes": [
            {"id": "minimumAscend", "label": "Early ascend"},
            {"id": "safeAscend", "label": "Safe ascend"},
            {"id": "optimalReset", "label": "Optimal ascend"},
        ],
        "defaults": {
            "pillar": "skills",
            "currentLevel": 1,
            "targetLevel": 100,
            "targetMode": "safeAscend",
            "discountPct": 0,
            "extraDropPct": 0,
            "dungeonLevel": 1,
            "clanTier": "B",
            "clanWinRate": 0.5,
            "rankedLeague": "gold",
            "rankBracket": "21-50",
            "includeRankedLeague": True,
            "includeMilestoneRewards": False,
            "currentResources": {
                "tickets": 0,
                "eggshells": 0,
                "clockwinders": 0,
            },
            "manualDailyIncome": {
                "tickets": 0,
                "eggshells": 0,
                "clockwinders": 0,
            },
        },
        "assumptionNotices": [
            "Skill tickets use a rounded per-key fitted curve from the 1-1, 9-4, and 15-6 anchors. Eggshells use a rounded per-key scale from the same anchors.",
            RESOURCE_MAPPING_NOTE,
        ],
        "sourceWorkbook": str(WORKBOOK_PATH),
    }


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(WORKBOOK_PATH, data_only=True)

    for spec in PILLARS:
        write_json(OUTPUT_DIR / f"{spec.name}.json", extract_progression_sheet(wb, spec))

    write_json(OUTPUT_DIR / "forge.json", extract_forge_progression(wb))
    write_json(OUTPUT_DIR / "ascensionTargets.json", extract_ascension_targets(wb))
    write_json(OUTPUT_DIR / "clanWarRewards.json", extract_clan_war_rewards(wb))
    write_json(OUTPUT_DIR / "rankedLeagueRewards.json", extract_ranked_league_rewards(wb))
    write_json(OUTPUT_DIR / "individualClanRewards.json", extract_individual_clan_rewards(wb))
    write_json(OUTPUT_DIR / "dungeonYieldConfig.json", build_dungeon_yields())
    write_json(OUTPUT_DIR / "appConfig.json", build_app_config())


if __name__ == "__main__":
    main()
